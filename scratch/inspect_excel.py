import openpyxl
import pandas as pd

excel_path = r"C:\Users\DELL\Desktop\Master_ AY Projects\trustmicro-credit\icare-group-member-onboarding-template.xlsx"
wb = openpyxl.load_workbook(excel_path, data_only=True)
print("Sheet names:", wb.sheetnames)

for sheet in wb.sheetnames:
    print(f"\n--- SHEET: {sheet} ---")
    df = pd.read_excel(excel_path, sheet_name=sheet)
    print("Columns:", df.columns.tolist())
    print("Row count:", len(df))
    print(df.head(5))
