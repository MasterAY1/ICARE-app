import openpyxl
import pandas as pd

def inspect_file(filename):
    print(f"\n==================================================")
    print(f"FILE: {filename}")
    print(f"==================================================")
    wb = openpyxl.load_workbook(filename, read_only=True)
    sheetnames = wb.sheetnames
    print(f"Sheets: {sheetnames}")
    for s in sheetnames:
        print(f"\n--- Sheet: {s} ---")
        df = pd.read_excel(filename, sheet_name=s)
        print(f"Shape: {df.shape}")
        print(f"Columns: {df.columns.tolist()[:15]}")
        print("First 2 rows:")
        print(df.head(2).to_string())

inspect_file("icare-group-member-onboarding-template.xlsx")
inspect_file("Co leder.xlsx")
inspect_file("Credit_Cash_Book_Ledger.xlsx")
inspect_file("clients.xlsx")
